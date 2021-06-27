import os

import time
from datetime import datetime, timezone, timedelta

import openpyxl
from openpyxl.worksheet import worksheet

from blivedm import GuardBuyMessage

class WorkbookMetadata:
    def __init__(self, filename: str, sheetname: str, head: list):
        self.filename = filename
        self.sheetname = sheetname
        self.head = head

guard_info = WorkbookMetadata(
    '大航海记录.xlsx', 
    '大航海记录', 
    ['弹幕ID', 'B站昵称', 'UID', '日期', '时间', '等级', '数量']
)

guard_bonus = WorkbookMetadata(
    '大航海福利.xlsx',
    '大航海福利',
    ['B站昵称', 'UID', '日期', '礼物', '消耗次数']
)

summary = WorkbookMetadata(
    '大航海信息汇总.xlsx',
    '大航海信息汇总',
    ['B站昵称', 'UID', '累计舰长次数', '已兑换', '剩余次数']
)

_level_to_text = {0: '非舰长', 1: '总督', 2: '提督', 3: '舰长'}
_level_to_num = {0: 0, 1: 100, 2: 10, 3: 1}

def _now():
    return datetime.now(timezone(timedelta(hours=8)))

def checkFileExist(filename: str):
    try:
        workbook = openpyxl.load_workbook(filename)
    except:
        workbook = openpyxl.Workbook()
        workbook.save(filename)
    return workbook

def checkSheetExist(filename: str, sheetname: str):
    workbook = checkFileExist(filename)
    try:
        worksheet = workbook[sheetname]
        return workbook, worksheet
    except:
        worksheet = workbook.create_sheet(sheetname)
        for record in [guard_info, guard_bonus, summary]:
            if filename == record.filename and sheetname == record.sheetname:
                for index, item in enumerate(record.head):
                    worksheet.cell(1, index + 1, item)
                workbook.save(filename)
                return workbook, worksheet
    raise 'Undefined filename or sheetname'

def saveGuardMessage(message: GuardBuyMessage):
    workbook, worksheet = checkSheetExist(guard_info.filename, guard_info.sheetname)
    worksheet.append([
        '%s-%s-%d-%d' % (message.uid, _now().strftime('%Y_%m_%d_%H_%M_%S'), message.start_time, message.end_time),
        message.username, 
        message.uid, 
        _now().strftime('%Y-%m-%d'), 
        _now().strftime('%H:%M:%S'), 
        _level_to_text[message.guard_level], 
        message.num
    ])
    workbook.save(guard_info.filename)

def getGuardInfo():
    # ['弹幕ID', 'B站昵称', 'UID', '日期', '时间', '等级', '数量']

    _, worksheet = checkSheetExist(guard_info.filename, guard_info.sheetname)
    accumulated_num = {}
    id_to_username = {}
    for row in list(worksheet.rows)[1:]:
        username = row[1]
        uid = row[2]
        plus_num = _level_to_num[row[5]] * row[6]

        id_to_username[uid] = username
        
        if uid not in accumulated_num:
            accumulated_num[uid] = 0
        accumulated_num = accumulated_num[uid] + plus_num

    return id_to_username, accumulated_num

def getGuardBonus():
    # ['B站昵称', 'UID', '日期', '礼物', '消耗次数']
    _, worksheet = checkSheetExist(guard_bonus.filename, guard_bonus.sheetname)
    accumulated_bonus = {}
    for row in list(worksheet.rows)[1:]:
        uid = row[1]
        plus_bonus = row[4]
        if uid not in accumulated_bonus:
            accumulated_bonus[uid] = 0
        accumulated_bonus[uid] = accumulated_bonus[uid] + plus_bonus
    return accumulated_bonus

def createSummary(message:saveGuardMessage):
    # ['B站昵称', 'UID', '累计舰长次数', '已兑换', '剩余次数']

    id_to_username, accumualted_num = getGuardInfo()
    accumualted_bonus = getGuardBonus()

    workbook, worksheet = checkSheetExist(guard_bonus.filename, guard_bonus.sheetname)
    for uid in accumualted_num:
        worksheet.append([
            id_to_username[uid],
            uid,
            accumualted_num[uid],
            accumualted_bonus[uid],
            accumualted_num[uid] - accumualted_bonus[uid]
        ])
    workbook.save(summary.filename)

def updateSummary(message: GuardBuyMessage):
    workbook, worksheet = checkSheetExist(guard_bonus.filename, guard_bonus.sheetname)

    uid = message.uid
    plus_num = _level_to_num[message.guard_level] * message.num

    # Scan all the items. If there is already a record for the user, then update it.
    for row in list(worksheet.rows):
        if row[1] == uid:
            row[2] = row[2] + plus_num
            row[4] = row[4] + plus_num
            return
    
    # If there is no record for the user, then create a new record.
    worksheet.append([
        message.username,
        uid,
        plus_num,
        0,
        plus_num
    ])
    workbook.save(summary.filename)


def updateGuardInfo(message: GuardBuyMessage):
    # Save raw information to the file
    saveGuardMessage(message)

    # Create or update the summary information
    if not os.path.isfile(summary.filename):
        createSummary(message)
    updateSummary(message)